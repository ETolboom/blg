import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from checks import CheckComplexity, CheckDetail, CheckFormInput, CheckResult
from rules.manager import RuleEvaluationSummary


class Assignment(BaseModel):
    # Set default value in case we want to onboard without a reference model
    reference_xml: str | None = ""


class GroupResultSummary(BaseModel):
    best_rule_id: str | None = None
    earned_points: float = 0.0
    rule_results: list[RuleEvaluationSummary] = []


class SubmissionCriterionResult(BaseModel):
    id: str
    score: float | None = None
    fulfilled: bool | None = None
    confidence: float = 0.0
    problematic_elements: list[str] = []
    inputs: list[CheckFormInput] = []
    group_result: GroupResultSummary | None = None
    detail: CheckDetail | None = None
    threshold_override: float | None = None
    ideal_threshold_override: float | None = None
    # Grader annotations split by audience: internal notes stay between graders,
    # feedback notes are intended for the student.
    internal_notes: str | None = None
    feedback_notes: str | None = None


class SubmissionResult(BaseModel):
    criteria: list[SubmissionCriterionResult] = []


class RubricCriterion(CheckResult):
    fulfilled: bool | None = None
    score: float | None = None
    default_points: float = 1.0
    group_result: GroupResultSummary | None = None
    supports_threshold: bool = False
    default_threshold: float | None = None
    default_ideal_threshold: float | None = None
    project_threshold: float | None = None
    project_ideal_threshold: float | None = None
    threshold_override: float | None = None
    ideal_threshold_override: float | None = None
    internal_notes: str | None = None
    feedback_notes: str | None = None
    threshold_label: str | None = None
    ideal_threshold_label: str | None = None
    threshold_hint: str | None = None


class CriterionDefinition(BaseModel):
    id: str
    name: str
    check_complexity: CheckComplexity
    description: str = ""
    default_points: float = 1.0
    inputs: list[CheckFormInput] = []
    supports_threshold: bool = False
    default_threshold: float | None = None
    default_ideal_threshold: float | None = None
    project_threshold: float | None = None
    project_ideal_threshold: float | None = None


class RubricDefinition(BaseModel):
    criteria: list[CriterionDefinition]
    assignment: Assignment | None = None

    def to_disk_json(self) -> str:
        """Serialize the rubric definition, excluding the reference XML."""
        return self.model_dump_json(exclude={"assignment": {"reference_xml"}})


class OnboardingRubric(BaseModel):
    assignment: Assignment
    checks: list[str] = []


class Rubric(BaseModel):
    criteria: list[RubricCriterion]
    assignment: Assignment | None = None

    def to_disk_json(self) -> str:
        """Serialize the rubric to JSON, excluding the reference XML."""
        return self.model_dump_json(exclude={"assignment": {"reference_xml"}})

    def to_excel_worksheet(
        self,
        workbook: Workbook,
        filename: str,
        include_threshold: bool = True,
        include_internal_notes: bool = True,
        include_feedback_notes: bool = True,
    ) -> None:
        """Write rubric criteria and scores as a formatted worksheet in the given
        workbook. ``include_threshold`` toggles the optional Threshold column;
        ``include_internal_notes``/``include_feedback_notes`` each toggle their own
        notes column (e.g. omit internal notes for a student-facing export)."""
        worksheet = workbook.create_sheet(filename)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="2E75B6", end_color="2E75B6", fill_type="solid"
        )
        criterion_font = Font(bold=True, size=11)
        criterion_fill = PatternFill(
            start_color="E8F1FF", end_color="E8F1FF", fill_type="solid"
        )
        border = Border(
            left=Side(border_style="thin", color="CCCCCC"),
            right=Side(border_style="thin", color="CCCCCC"),
            top=Side(border_style="thin", color="CCCCCC"),
            bottom=Side(border_style="thin", color="CCCCCC"),
        )
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        center_alignment = Alignment(horizontal="center", vertical="center")

        def calculate_points(criterion):
            if not criterion.fulfilled:
                return 0.0
            elif criterion.score is not None:
                return max(0.0, criterion.score)
            else:
                return max(0.0, criterion.default_points)

        def _line(label, override, default) -> str | None:
            """One threshold line, emitted only when the override deviates from
            the default. An unset (or default-equal) threshold implies the
            default, so it is omitted to keep the column free of noise."""
            if override is not None and override != default:
                return f"{label} {override} (default {default})"
            return None

        def threshold_cell(criterion) -> str:
            """Make deviations traceable: list only the cut-offs (min and/or
            ideal) the grader overrode away from the effective default; criteria
            graded at the effective thresholds leave this column blank. The
            effective default is the project threshold when set, otherwise the
            global default."""
            if not criterion.supports_threshold:
                return ""
            eff_min = (
                criterion.project_threshold
                if criterion.project_threshold is not None
                else criterion.default_threshold
            )
            eff_ideal = (
                criterion.project_ideal_threshold
                if criterion.project_ideal_threshold is not None
                else criterion.default_ideal_threshold
            )
            lines = [
                _line("min", criterion.threshold_override, eff_min),
                _line("ideal", criterion.ideal_threshold_override, eff_ideal),
            ]
            return "\n".join(line for line in lines if line)

        # Build the column layout, dropping optional columns when toggled off so
        # the indices/widths shift to keep the sheet gap-free.
        columns: list[tuple[str, int, bool]] = [
            ("Criterion", 25, False),
            ("Description", 45, False),
            ("Points", 10, True),
        ]
        if include_threshold:
            columns.append(("Threshold", 22, False))
        if include_internal_notes:
            columns.append(("Internal notes", 40, False))
        if include_feedback_notes:
            columns.append(("Feedback notes", 40, False))

        # Add headers
        for col, (header, _width, _centered) in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        def cell_value(criterion, header: str):
            if header == "Criterion":
                return criterion.name
            if header == "Description":
                return criterion.description
            if header == "Points":
                return calculate_points(criterion)
            if header == "Threshold":
                return threshold_cell(criterion)
            if header == "Internal notes":
                return criterion.internal_notes or ""
            if header == "Feedback notes":
                return criterion.feedback_notes or ""
            return ""

        # Add criteria data
        current_row = 2
        for criterion in self.criteria:
            for col, (header, _width, centered) in enumerate(columns, 1):
                cell = worksheet.cell(
                    row=current_row, column=col, value=cell_value(criterion, header)
                )
                cell.border = border
                cell.font = criterion_font
                cell.fill = criterion_fill
                cell.alignment = center_alignment if centered else wrap_alignment

            current_row += 1

        # Set column widths
        for col, (_header, width, _centered) in enumerate(columns, 1):
            worksheet.column_dimensions[get_column_letter(col)].width = width

    def to_excel(
        self,
        filename: str,
        include_threshold: bool = True,
        include_internal_notes: bool = True,
        include_feedback_notes: bool = True,
    ) -> bytes:
        """Export the rubric to an Excel workbook and return its raw bytes."""
        excel_buffer = io.BytesIO()
        workbook = Workbook()

        self.to_excel_worksheet(
            workbook,
            filename,
            include_threshold=include_threshold,
            include_internal_notes=include_internal_notes,
            include_feedback_notes=include_feedback_notes,
        )

        # Remove default sheet if it exists
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
