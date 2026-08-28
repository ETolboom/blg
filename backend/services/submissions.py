import glob
import io
import json
import os

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook

from rubric import (
    RubricDefinition,
    Rubric,
    RubricCriterion,
    SubmissionCriterionResult,
    SubmissionResult,
)
from checks import ThresholdMeta
from demo_mode import demo_mode_enabled
from utils import safe_join

# Maximum accepted size for an uploaded .bpmn submission.
MAX_SUBMISSION_SIZE_MB = 10
MAX_SUBMISSION_SIZE = MAX_SUBMISSION_SIZE_MB * 1024 * 1024

# Sentinel "filename" that refers to the reference model rather than a submission.
# Its evaluation is stored as <project>/reference.bpmn.json.
REFERENCE_FILENAME = "Reference"


class SubmissionService:
    def __init__(self, base_path: str, rubric: RubricDefinition | None, check_registry=None):
        """Initialize the service with the data directory path and the rubric
        definition. ``check_registry`` (optional) lets composition derive each
        criterion's threshold metadata from the live check classes, so it is
        correct even for rubrics authored before threshold support existed."""
        self.base_path = base_path
        self.submissions_path = os.path.join(base_path, "submissions")
        self.rubric = rubric
        self.check_registry = check_registry

    def list_submissions(self) -> list[dict]:
        """Return filename, display name and analyzed flag for every .bpmn file
        in the submissions directory. ``analyzed`` is True when a per-model
        evaluation exists — only analyzed submissions can be exported."""
        # In demo mode the data root is read-only; the directory is expected to
        # exist, and a missing one fails on the listdir below with a clear error.
        if not demo_mode_enabled():
            os.makedirs(self.submissions_path, exist_ok=True)
        return [
            {
                "filename": f,
                "name": f.removesuffix(".bpmn"),
                "analyzed": os.path.exists(self._eval_path(f)),
            }
            for f in os.listdir(self.submissions_path)
            if f.endswith(".bpmn")
        ]

    def get_submission_xml(self, filename: str) -> str:
        """Return raw BPMN XML for a submission; pass REFERENCE_FILENAME for the reference model."""
        if filename == REFERENCE_FILENAME:
            if (
                self.rubric
                and self.rubric.assignment
                and self.rubric.assignment.reference_xml
            ):
                return self.rubric.assignment.reference_xml
            else:
                raise HTTPException(status_code=404, detail="Reference XML not found")

        path = safe_join(self.submissions_path, filename)
        with open(path) as f:
            return f.read()

    def _eval_path(self, filename: str) -> str:
        """Path of the per-model evaluation file.

        The reference is treated as just another evaluated model, stored as
        ``reference.bpmn.json`` next to the reference; submissions live under
        ``submissions/<filename>.json``.
        """
        if filename == REFERENCE_FILENAME:
            return os.path.join(self.base_path, "reference.bpmn.json")
        return safe_join(self.submissions_path, filename + ".json")

    def _read_eval(self, filename: str) -> SubmissionResult | None:
        """Read a per-model evaluation file and return a SubmissionResult."""
        path = self._eval_path(filename)
        if not os.path.exists(path):
            return None

        with open(path, encoding="utf-8") as f:
            return SubmissionResult.model_validate(json.load(f))

    def write_eval(self, filename: str, result: SubmissionResult) -> None:
        """Persist a per-model evaluation file."""
        with open(self._eval_path(filename), "w", encoding="utf-8") as f:
            f.write(result.model_dump_json())

    def compose_rubric(
        self, filename: str, result: SubmissionResult | None = None
    ) -> Rubric:
        """Compose a full Rubric by merging the ground-truth definition with a
        model's evaluation results. A missing evaluation yields ungraded
        criteria (fulfilled=None) rather than an error.

        Pass ``result`` to compose from an in-memory evaluation instead of the
        persisted file (used where the result is deliberately not written)."""
        if self.rubric is None:
            raise HTTPException(status_code=404, detail="No rubric loaded")

        if result is None:
            result = self._read_eval(filename) or SubmissionResult()
        result_map = {cr.id: cr for cr in result.criteria}

        composed: list[RubricCriterion] = []
        for ref in self.rubric.criteria:
            sr = result_map.get(ref.id)
            # Prefer live check metadata (works for older rubrics that predate
            # threshold support); fall back to whatever the definition stored.
            if self.check_registry is not None:
                meta = self.check_registry.threshold_meta(ref.id)
            else:
                meta = ThresholdMeta(
                    supports=ref.supports_threshold,
                    default_threshold=ref.default_threshold,
                    default_ideal_threshold=ref.default_ideal_threshold,
                )
            composed.append(
                RubricCriterion(
                    id=ref.id,
                    name=ref.name,
                    description=ref.description,
                    check_complexity=ref.check_complexity,
                    inputs=ref.inputs,
                    default_points=ref.default_points,
                    fulfilled=sr.fulfilled if sr else None,
                    score=sr.score if sr else None,
                    confidence=sr.confidence if sr else 0.0,
                    problematic_elements=sr.problematic_elements if sr else [],
                    group_result=sr.group_result if sr else None,
                    detail=sr.detail if sr else None,
                    counter_example=sr.counter_example if sr else None,
                    # Threshold support/defaults/labels derived above; the
                    # overrides and notes are per-submission deviations from the
                    # evaluation.
                    supports_threshold=meta.supports,
                    default_threshold=meta.default_threshold,
                    default_ideal_threshold=meta.default_ideal_threshold,
                    project_threshold=ref.project_threshold,
                    project_ideal_threshold=ref.project_ideal_threshold,
                    threshold_label=meta.threshold_label,
                    ideal_threshold_label=meta.ideal_threshold_label,
                    threshold_hint=meta.threshold_hint,
                    threshold_override=sr.threshold_override if sr else None,
                    ideal_threshold_override=(
                        sr.ideal_threshold_override if sr else None
                    ),
                    internal_notes=sr.internal_notes if sr else None,
                    feedback_notes=sr.feedback_notes if sr else None,
                )
            )

        # Carry the project's assignment (incl. reference_xml, populated at load)
        # so the grading view can render the reference model.
        return Rubric(criteria=composed, assignment=self.rubric.assignment)

    def invalidate_all_results(self) -> None:
        """Delete all cached .bpmn.json result files."""
        for path in glob.glob(os.path.join(self.submissions_path, "*.bpmn.json")):
            os.remove(path)

    def get_submission_rubric(self, filename: str) -> Rubric:
        """Return the composed rubric for the given submission filename."""
        return self.compose_rubric(filename)

    async def upload_submissions(self, files: list[UploadFile]) -> list[dict]:
        """Save uploaded BPMN files to the submissions directory and return their metadata."""
        os.makedirs(self.submissions_path, exist_ok=True)

        uploaded = []
        for file in files:
            if not file.filename or not file.filename.endswith(".bpmn"):
                raise HTTPException(
                    status_code=400,
                    detail=f"'{file.filename}' is not a .bpmn file",
                )

            content = await file.read()
            if len(content) > MAX_SUBMISSION_SIZE:
                raise HTTPException(
                    status_code=400,
                    detail=f"'{file.filename}' exceeds the {MAX_SUBMISSION_SIZE_MB}MB limit",
                )

            dest = safe_join(self.submissions_path, file.filename)
            with open(dest, "wb") as f:
                f.write(content)

            uploaded.append(
                {"filename": file.filename, "name": file.filename.removesuffix(".bpmn")}
            )

        return uploaded

    def update_submission_criteria(
        self, filename: str, criteria: list[SubmissionCriterionResult]
    ) -> None:
        """Merge updated criterion results into the stored submission result file."""
        if not os.path.exists(self._eval_path(filename)):
            raise HTTPException(status_code=404, detail="Submission not found")

        # Read existing result (handles backward compat)
        result = self._read_eval(filename) or SubmissionResult()

        # Build map from existing criteria, then overlay updates. Preserve an
        # existing group breakdown when a manual update doesn't carry one.
        existing_map = {cr.id: cr for cr in result.criteria}
        for updated in criteria:
            prev = existing_map.get(updated.id)
            if updated.group_result is None and prev is not None and prev.group_result:
                updated.group_result = prev.group_result
            existing_map[updated.id] = updated

        result.criteria = list(existing_map.values())
        self.write_eval(filename, result)

    def apply_criterion_result(
        self, filename: str, result: SubmissionCriterionResult
    ) -> None:
        """Overlay a single freshly-evaluated criterion onto the stored eval,
        preserving the grader's existing notes for that criterion (a re-grade
        recomputes scoring, not the manual annotations)."""
        existing = self._read_eval(filename) or SubmissionResult()
        by_id = {cr.id: cr for cr in existing.criteria}
        prev = by_id.get(result.id)
        if prev is not None:
            result.internal_notes = prev.internal_notes
            result.feedback_notes = prev.feedback_notes
        by_id[result.id] = result
        existing.criteria = list(by_id.values())
        self.write_eval(filename, existing)

    def regrade_inheriting_submissions(
        self, crit, registry, rule_manager
    ) -> None:
        """Re-grade ``crit`` for every analyzed submission that inherits its
        threshold, picking up a freshly-changed project threshold.

        Submissions carrying their own override for the criterion are left
        untouched (a submission override wins over the project level); the grader
        note is preserved via ``apply_criterion_result``. A criterion counts as
        overridden when either its min or ideal override is set; both notes are
        preserved via ``apply_criterion_result``.
        """
        # Imported here to avoid a module-level import cycle with the router layer
        # that wires evaluation and this service together.
        from services.evaluation import evaluate_criterion

        for entry in self.list_submissions():
            if not entry["analyzed"]:
                continue
            filename = entry["filename"]
            existing = self._read_eval(filename)
            if existing is None:
                continue
            sr = next((c for c in existing.criteria if c.id == crit.id), None)
            if sr is not None and (
                sr.threshold_override is not None
                or sr.ideal_threshold_override is not None
            ):
                continue

            model_xml = self.get_submission_xml(filename)
            manager = registry.create_manager(model_xml)
            result = evaluate_criterion(crit, manager, model_xml, rule_manager)
            self.apply_criterion_result(filename, result)

    def export_submissions(
        self,
        filenames: list[str],
        include_thresholds: bool = True,
        include_internal_notes: bool = True,
        include_feedback_notes: bool = True,
    ) -> bytes:
        """Export the graded rubrics for the given submissions into one workbook,
        a worksheet per submission. ``include_thresholds`` toggles the Threshold
        column; ``include_internal_notes``/``include_feedback_notes`` each toggle
        their own notes column."""
        if not filenames:
            raise HTTPException(
                status_code=400, detail="No submissions selected for export"
            )

        workbook = Workbook()
        for filename in filenames:
            self.compose_rubric(filename).to_excel_worksheet(
                workbook,
                filename.removesuffix(".bpmn"),
                include_threshold=include_thresholds,
                include_internal_notes=include_internal_notes,
                include_feedback_notes=include_feedback_notes,
            )

        # Drop openpyxl's empty default sheet — but only once we've added our own,
        # since a workbook must keep at least one sheet.
        if len(workbook.sheetnames) > 1 and "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
